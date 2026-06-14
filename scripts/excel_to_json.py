#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
excel_to_json.py

用途：
    解析 Excel 表格数据，导出 JSON。
    - 表头默认在第 2 行
    - 支持提取嵌入在单元格里的图片，保存到本地目录
    - JSON 中只写图片相对路径
    - 默认字段映射：
        题目ID   -> id
        题目原文 -> image_path
        答案     -> reference_answer
        题目出处 -> source
        所属学段 -> grade
        数学分支 -> math_type
        题目标签 -> tag
        难度等级 -> difficulty

安装依赖：
    pip install openpyxl

示例：
    python3 excel_to_json.py 数字人数学场景测试.xlsx -o output.json --sheet "Qwen3-32B-COT-EN-0612"

说明：
    1. 如果某个字段所在单元格有图片，脚本会优先把图片保存后写入图片相对路径。
    2. 如果一个单元格有多张图片，该 JSON 字段会写成路径列表。
    3. 如果单元格没有图片，则写入单元格文本/数值。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_FIELD_MAPPING = {
    "题目ID": "id",
    "题目原文": "image_path",
    "答案": "reference_answer",
    "题目出处": "source",
    "所属学段": "grade",
    "数学分支": "math_type",
    "题目标签": "tag",
    "难度等级": "difficulty",
}


def safe_name(value: Any, default: str = "unknown") -> str:
    """把任意字符串转成适合作为文件名的一部分。"""
    text = str(value).strip() if value is not None else default
    text = re.sub(r"[\\/:*?\"<>|\s]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or default


def normalize_cell_value(value: Any) -> Any:
    """把 Excel 单元格值转成适合 JSON 的值。"""
    if value is None:
        return None

    # 日期、时间等对象通常有 isoformat 方法。
    if hasattr(value, "isoformat"):
        return value.isoformat()

    return value


def get_image_start_cell(image: Any) -> Tuple[int, int]:
    """
    获取图片左上角锚定单元格，返回 1-based 的 (row, col)。

    openpyxl 内部 anchor._from 的 row/col 是 0-based。
    """
    anchor = image.anchor
    if not hasattr(anchor, "_from"):
        raise ValueError("图片缺少 anchor._from 信息，无法定位单元格")

    return anchor._from.row + 1, anchor._from.col + 1


def guess_image_ext(image: Any, image_bytes: bytes) -> str:
    """尽量可靠地判断图片扩展名。"""
    fmt = (getattr(image, "format", "") or "").lower().strip(".")
    if fmt in {"jpeg", "jpg"}:
        return "jpg"
    if fmt in {"png", "gif", "bmp", "webp"}:
        return fmt

    # 兜底：根据文件头猜测。
    if image_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if image_bytes.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if image_bytes.startswith(b"GIF87a") or image_bytes.startswith(b"GIF89a"):
        return "gif"
    if image_bytes.startswith(b"RIFF") and b"WEBP" in image_bytes[:16]:
        return "webp"

    return "bin"


def build_image_index(
    ws: Any,
    image_output_dir: Path,
    json_base_dir: Path,
    sheet_name: str,
    header_by_col: Dict[int, str],
) -> Dict[Tuple[int, int], List[str]]:
    """
    保存工作表中的图片，并建立：
        (row, col) -> [relative_image_path, ...]
    """
    image_output_dir.mkdir(parents=True, exist_ok=True)

    images_by_cell: Dict[Tuple[int, int], List[str]] = defaultdict(list)
    per_cell_count: Dict[Tuple[int, int], int] = defaultdict(int)

    for image in getattr(ws, "_images", []):
        try:
            row, col = get_image_start_cell(image)
            image_bytes = image._data()
        except Exception as exc:
            print(f"[WARN] 跳过一张无法读取或定位的图片：{exc}", file=sys.stderr)
            continue

        per_cell_count[(row, col)] += 1
        idx = per_cell_count[(row, col)]

        col_name = header_by_col.get(col, get_column_letter(col))
        ext = guess_image_ext(image, image_bytes)

        filename = f"{safe_name(sheet_name)}_r{row}_c{col}_{safe_name(col_name)}_{idx}.{ext}"
        image_path = image_output_dir / filename

        with image_path.open("wb") as f:
            f.write(image_bytes)

        # JSON 里写相对 output.json 所在目录的路径。
        rel_path = image_path.relative_to(json_base_dir).as_posix()
        images_by_cell[(row, col)].append(rel_path)

    return images_by_cell


def choose_field_value(cell_value: Any, image_paths: List[str]) -> Any:
    """
    字段取值规则：
    - 有图片：返回图片相对路径；多张图片返回路径列表
    - 无图片：返回单元格文本/数值
    """
    if image_paths:
        return image_paths[0] if len(image_paths) == 1 else image_paths
    return normalize_cell_value(cell_value)


def parse_sheet(
    xlsx_path: Path,
    sheet_name: Optional[str],
    output_json_path: Path,
    image_dir_name: str,
    header_row: int,
    data_start_row: Optional[int],
    field_mapping: Dict[str, str],
    skip_empty_id: bool = True,
) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    if data_start_row is None:
        data_start_row = header_row + 1

    output_json_path.parent.mkdir(parents=True, exist_ok=True)
    image_output_dir = output_json_path.parent / image_dir_name

    # 读取表头：表头名 -> 列号
    header_to_col: Dict[str, int] = {}
    header_by_col: Dict[int, str] = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue

        header = str(value).strip()
        if not header:
            continue

        # 遇到重复表头时，保留第一次出现的列，避免后面的重复列覆盖前面的列。
        if header not in header_to_col:
            header_to_col[header] = col
        header_by_col[col] = header

    missing_headers = [name for name in field_mapping if name not in header_to_col]
    if missing_headers:
        raise ValueError(
            "表格缺少以下表头："
            + ", ".join(missing_headers)
            + f"\n当前工作表：{ws.title}\n实际读取到的表头：{list(header_to_col.keys())}"
        )

    images_by_cell = build_image_index(
        ws=ws,
        image_output_dir=image_output_dir,
        json_base_dir=output_json_path.parent,
        sheet_name=ws.title,
        header_by_col=header_by_col,
    )

    records: List[Dict[str, Any]] = []

    for row in range(data_start_row, ws.max_row + 1):
        record: Dict[str, Any] = {}

        for excel_header, json_field in field_mapping.items():
            col = header_to_col[excel_header]
            cell = ws.cell(row=row, column=col)
            image_paths = images_by_cell.get((row, col), [])
            record[json_field] = choose_field_value(cell.value, image_paths)

        # 默认跳过没有题目ID的行，避免空行进入 JSON。
        if skip_empty_id and not record.get("id"):
            continue

        # 如果整行都是空，也跳过。
        if all(value in (None, "") or value == [] for value in record.values()):
            continue

        records.append(record)

    return records


def load_mapping(mapping_json: Optional[str]) -> Dict[str, str]:
    """支持从命令行传入自定义字段映射 JSON。"""
    if not mapping_json:
        return DEFAULT_FIELD_MAPPING

    try:
        loaded = json.loads(mapping_json)
    except json.JSONDecodeError as exc:
        raise ValueError(f"--mapping 不是合法 JSON：{exc}") from exc

    if not isinstance(loaded, dict):
        raise ValueError("--mapping 必须是 JSON 对象，例如：'{\"题目ID\":\"id\"}'")

    return {str(k): str(v) for k, v in loaded.items()}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="解析 Excel 表格，提取嵌入图片，并导出 JSON。"
    )
    parser.add_argument("xlsx", help="输入 Excel 文件路径，例如：input.xlsx")
    parser.add_argument("-o", "--output", default="output.json", help="输出 JSON 路径，默认：output.json")
    parser.add_argument("--sheet", default=None, help="工作表名称；不传则使用 Excel 当前激活工作表")
    parser.add_argument("--header-row", type=int, default=2, help="表头所在行，默认：2")
    parser.add_argument("--data-start-row", type=int, default=None, help="数据起始行，默认：header-row + 1")
    parser.add_argument("--image-dir", default="images", help="图片保存目录，相对 output.json 所在目录，默认：images")
    parser.add_argument(
        "--mapping",
        default=None,
        help=(
            "自定义字段映射 JSON 字符串。"
            "例如：'{\"题目ID\":\"id\",\"题目原文\":\"image_path\"}'"
        ),
    )
    parser.add_argument(
        "--keep-empty-id",
        action="store_true",
        help="保留没有题目ID的行；默认会跳过没有题目ID的行",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    output_json_path = Path(args.output).expanduser().resolve()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{xlsx_path}")

    field_mapping = load_mapping(args.mapping)

    records = parse_sheet(
        xlsx_path=xlsx_path,
        sheet_name=args.sheet,
        output_json_path=output_json_path,
        image_dir_name=args.image_dir,
        header_row=args.header_row,
        data_start_row=args.data_start_row,
        field_mapping=field_mapping,
        skip_empty_id=not args.keep_empty_id,
    )

    with output_json_path.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f"[OK] 导出完成：{output_json_path}")
    print(f"[OK] 记录数量：{len(records)}")
    print(f"[OK] 图片目录：{output_json_path.parent / args.image_dir}")


if __name__ == "__main__":
    main()
